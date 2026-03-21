#!/usr/bin/env python3
"""
Add / refresh a tab "Dashboard_Snapshot" on the ESA sales workbook with KPIs from the live dashboard API.

Usage:
  python snapshot_dashboard_tab.py [workbook.xlsx]
  python snapshot_dashboard_tab.py --range 7d
  python snapshot_dashboard_tab.py --url https://esa-marketing-sales-dashboard.vercel.app/api/data

Default workbook: ~/Downloads/Untitled spreadsheet.xlsx
Requires: openpyxl (same venv as clean_sales_sheet: .venv-sheet)
"""

from __future__ import annotations

import argparse
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

SNAPSHOT_TAB = "Dashboard_Snapshot"
DEFAULT_API = "https://esa-marketing-sales-dashboard.vercel.app/api/data"
DEFAULT_BOOK = Path.home() / "Downloads" / "Untitled spreadsheet.xlsx"


def fetch_json(url: str, timeout: int = 90) -> dict:
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def flatten_metrics(d: dict) -> list[tuple[str, str, str]]:
    """Human label, display value, how we got it (keep in sync with main.js buildSnapshotRows)."""
    sc = d.get("statusCounts") or {}
    lf = d.get("leadFormMarketing") or {}
    rm = d.get("revenueModel") or {}
    lines: list[tuple[str, str, str]] = []
    ad_src = d.get("adSource") or "Meta Marketing API (or empty if META_ACCESS_TOKEN / account not set)"

    def s(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            return str(v) if v == int(v) else f"{v:.2f}"
        return str(v)

    ghl_c = (
        "GHL REST API · contacts with ESA tags · dateAdded (or createdAt) inside selected range"
    )
    ghl_opp_rev = (
        "Google Sheets API · SHEETS_REVENUE_MODE=replace · rows whose date column falls in range"
        if rm.get("source") == "google_sheets"
        else "GHL REST API · pipeline opportunities · "
        + str(rm.get("attribution") or "Closed Won; sum monetaryValue")
    )

    def tri(a: str, b, c: str) -> None:
        lines.append((a, s(b), c))

    tri("Fetched (UTC)", d.get("fetchedAt"), "Server · when /api/data finished building this JSON")
    tri("Dashboard date range", d.get("dateRange"), "Server · computed window from ?range= / ?start= / ?end=")
    tri("API range param", d.get("_queryRange"), "CLI/query param passed to this script")
    lines.append(("", "", ""))

    tri("Revenue ($)", d.get("revenue"), ghl_opp_rev)
    tri("Ad spend ($)", d.get("adSpend"), ad_src)
    tri(
        "Upfront ROAS (x)",
        d.get("upfrontRoas"),
        "Derived · Revenue ÷ Ad spend (same date range)",
    )
    tri(
        "Revenue source",
        rm.get("source", ""),
        "API field revenueModel.source · which system owns the revenue number above",
    )
    closed_how = (
        "Google Sheets · deal row count in range (when sheet replaces GHL revenue)"
        if rm.get("source") == "google_sheets"
        else "GHL REST API · opportunities in Closed Won with lastStatusChangeAt in range"
    )
    tri("Closed won deals (opp count)", sc.get("closedWonDeals"), closed_how)
    tri("AOV ($)", d.get("aov"), "Derived · Revenue ÷ closed-won deal count")
    tri("CPA ($)", d.get("cpa"), "Derived · Ad spend ÷ closed-won deal count")
    lines.append(("", "", ""))

    tri("GHL leads (in range)", d.get("leads"), ghl_c + " · counted as leads")
    tri("Booked calls", d.get("bookedCalls"), ghl_c + " · status-booked tag")
    tri("Lead → book %", d.get("leadBookPct"), "Derived · booked ÷ leads")
    tri("Showed", sc.get("showed"), ghl_c + " · status-showed tag")
    tri("Show rate %", d.get("showRate"), "Derived · showed ÷ booked")
    tri("Offers made", sc.get("offered"), ghl_c + " · status-offer-made tag")
    tri("Closed won (tags)", sc.get("closedWon"), ghl_c + " · status-closed-won tag (not opp $)")
    tri("Meta Results (sum campaigns)", d.get("metaLeads"), ad_src + " · sum of campaign results / leads")
    tri("Meta impressions", d.get("metaImpressions"), ad_src + " · account insights")
    tri("Meta clicks", d.get("metaClicks"), ad_src + " · account insights")
    tri("CPL ($)", d.get("cpl"), "Derived · ad spend ÷ GHL leads (in range)")
    tri("Cost / booking ($)", d.get("costPerBooking"), "Derived · ad spend ÷ booked calls")
    lines.append(("", "", ""))

    tri(
        "SS Lead Form · Meta campaign results",
        lf.get("metaSSLeadFormResults"),
        "Meta API · campaigns tagged for SS lead form in dashboard code · results field",
    )
    tri(
        "SS Lead Form · GHL cohort count",
        lf.get("total"),
        "GHL contacts in range · src-fb-lead-form-ss / Meta lead-form tag rules (api/data.js)",
    )
    tri("SS · Booked", lf.get("booked"), "Same SS cohort · subset with status-booked")
    tri(
        "Canonical Meta lead form GHL tag",
        lf.get("canonicalGhlMetaLeadFormTag", ""),
        "Constant in api/data.js · used to match Meta lead form to GHL tags",
    )
    lines.append(("", "", ""))

    tri(
        "Pipeline value (open opps $)",
        d.get("pipelineValue"),
        "GHL REST API · all opps in Brian & Diamond pipeline · sum monetaryValue by stage "
        "(current snapshot; not filtered by header date range)",
    )
    tri("Contacts in CRM (total)", d.get("totalContacts"), "GHL REST API · full contact sync for this location")
    tri("Contacts in date range", d.get("contactsInRange"), ghl_c + " · rows after date filter")

    sr = d.get("sheetRevenue") or {}
    if sr.get("mode") and sr.get("mode") != "off":
        lines.append(("", "", ""))
        tri("Sheet revenue mode", sr.get("mode"), "Env SHEETS_REVENUE_MODE · server-side toggle")
        tri(
            "Sheet revenue total",
            sr.get("totalRevenue"),
            "Google Sheets API · service account · spreadsheet in env (see README)",
        )
        if sr.get("warning"):
            tri(
                "Sheet warning",
                sr.get("warning"),
                "Sheet fetch or parse message · check Vercel env / sheet access",
            )

    return lines


def main():
    p = argparse.ArgumentParser(description="Add Dashboard_Snapshot tab from live API")
    p.add_argument("workbook", nargs="?", type=Path, default=DEFAULT_BOOK, help="Path to .xlsx")
    p.add_argument("--range", default="30d", help="Query range: 7d, 30d, 90d, custom (use with --start/--end)")
    p.add_argument("--start", default="", help="Custom range start YYYY-MM-DD")
    p.add_argument("--end", default="", help="Custom range end YYYY-MM-DD")
    p.add_argument("--url", default=DEFAULT_API, help="Base API URL (without query string)")
    args = p.parse_args()

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        sys.exit(1)

    q = f"?range={args.range}"
    if args.range == "custom" and args.start:
        q += f"&start={args.start}"
        if args.end:
            q += f"&end={args.end}"

    full_url = args.url.split("?")[0].rstrip("/") + q
    print(f"GET {full_url}")

    try:
        data = fetch_json(full_url)
        data["_queryRange"] = args.range
    except urllib.error.HTTPError as e:
        print(f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')[:500]}", file=sys.stderr)
        sys.exit(1)
    except urllib.error.URLError as e:
        print(f"Request failed: {e}", file=sys.stderr)
        sys.exit(1)

    if "error" in data:
        print(f"API error: {data.get('error')}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(args.workbook, data_only=False)
    if SNAPSHOT_TAB in wb.sheetnames:
        del wb[SNAPSHOT_TAB]
    ws = wb.create_sheet(SNAPSHOT_TAB)

    title_fill = PatternFill("solid", fgColor="DDEBF7")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    h = ws.cell(row=1, column=1, value="ESA dashboard snapshot — re-run: snapshot_dashboard_tab.py")
    h.font = Font(bold=True, size=12)
    h.fill = title_fill
    h.alignment = Alignment(wrap_text=True)

    ws.cell(row=2, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=2, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=2, column=3, value="How we got it").font = Font(bold=True)

    data["_queryRange"] = args.range
    rows = flatten_metrics(data)
    for i, (a, b, c) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value=c).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 64
    wb.save(args.workbook)
    print(f"Updated tab '{SNAPSHOT_TAB}' in {args.workbook} ({len(rows)} metric rows)")


if __name__ == "__main__":
    main()
