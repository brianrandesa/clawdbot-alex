#!/usr/bin/env python3
"""
ESA sales workbook helper:
- Rebuilds tab "Clean_All" from Sheet1 (clean dates, text, $).
- Ensures tab "Enter_Sales" exists (manual entry; replaces Google Forms flow).

Usage: python clean_sales_sheet.py [path/to/file.xlsx]
Default: ~/Downloads/Untitled spreadsheet.xlsx

Tab order: Sheet1 | Clean_All | Enter_Sales
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_EPOCH = datetime(1899, 12, 30)

CLEAN_HEADERS = [
    "Fathom_1",
    "Fathom_2",
    "Date_Created",
    "First_Sales_Call_Date",
    "Payment_Date",
    "Client_Or_Event",
    "Product",
    "Email",
    "Phone",
    "Amount_Paid",
    "Amount_Owed",
    "Setter",
    "Setter_5pct",
    "Closer",
    "Closer_Commission",
    "Lead_Source",
    "Campaign",
    "Adset",
    "Ad",
]

COLUMN_WIDTHS = [28, 28, 12, 12, 12, 36, 22, 28, 16, 12, 12, 12, 12, 14, 14, 18, 24, 24, 28]

ENTER_SALES_NAME = "Enter_Sales"

ENTER_INSTRUCTIONS = (
    "ENTER NEW SALES HERE (one row per payment). Required: Payment_Date, Client_Or_Event, Amount_Paid. "
    "Payment plans: new row per payment; Amount_Owed = balance still due after this payment. "
    "Optional: Fathom link, attribution, setter/closer. To refresh Clean_All: copy new rows onto Sheet1 "
    "(append under existing data), save, then run clean_sales_sheet.py again."
)


def excel_serial_to_date(n: float | int) -> datetime | None:
    if n is None:
        return None
    try:
        f = float(n)
    except (TypeError, ValueError):
        return None
    if not (20000 < f < 80000):
        return None
    return EXCEL_EPOCH + timedelta(days=int(f))


def clean_str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_money(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean_str(v).replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def is_probable_date_serial(v) -> bool:
    try:
        f = float(v)
        return 20000 < f < 80000
    except (TypeError, ValueError):
        return False


def apply_column_widths(ws) -> None:
    for i, w in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def ensure_enter_sales_sheet(wb) -> str:
    """Create Enter_Sales at tab position 3 if missing. Never delete user rows if sheet exists."""
    if ENTER_SALES_NAME in wb.sheetnames:
        return f"Tab '{ENTER_SALES_NAME}' already present (left your rows as-is)."

    # Third tab: index 2 after Sheet1 (0) and Clean_All (1)
    ws = wb.create_sheet(ENTER_SALES_NAME, 2)
    header_fill = PatternFill("solid", fgColor="E2EFDA")
    for col, h in enumerate(CLEAN_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(CLEAN_HEADERS))
    tip = ws.cell(row=2, column=1, value=ENTER_INSTRUCTIONS)
    tip.font = Font(italic=True, size=10, color="444444")
    tip.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 48

    apply_column_widths(ws)
    ws.freeze_panes = "A3"

    # Light grid from row 3 down (room to type)
    for r in range(3, 502):
        for c in range(1, len(CLEAN_HEADERS) + 1):
            cell = ws.cell(row=r, column=c, value=None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return f"Created tab '{ENTER_SALES_NAME}' (start typing in row 3)."


def main():
    default_path = Path.home() / "Downloads" / "Untitled spreadsheet.xlsx"
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not src.exists():
        print(f"File not found: {src}", file=sys.stderr)
        sys.exit(1)

    # Keep formulas on Sheet1 when saving
    wb = load_workbook(src, data_only=False)
    ws = wb.active

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        print("Empty sheet", file=sys.stderr)
        sys.exit(1)

    max_cols = max(len(r) for r in raw_rows if r)

    cleaned_data = []
    for row in raw_rows[1:]:
        if not row:
            continue
        cells = list(row) + [None] * (max_cols - len(row))
        if not any(c is not None and str(c).strip() for c in cells[:19]):
            continue
        name = clean_str(cells[5]) if len(cells) > 5 else ""
        email = clean_str(cells[7]) if len(cells) > 7 else ""
        ap = clean_money(cells[9]) if len(cells) > 9 else None
        if not name and not email and (ap is None or ap == 0):
            continue

        out = []
        for i in range(19):
            v = cells[i] if i < len(cells) else None
            if i in (9, 10):
                out.append(clean_money(v))
            elif i in (2, 3, 4):
                if isinstance(v, datetime):
                    out.append(v.date())
                elif isinstance(v, date):
                    out.append(v)
                elif v is not None and is_probable_date_serial(v):
                    d = excel_serial_to_date(v)
                    out.append(d.date() if d else "")
                else:
                    out.append(clean_str(v) if v is not None else "")
            elif i in (0, 1, 5, 6, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18):
                out.append(clean_str(v) if v is not None else "")
            else:
                out.append(clean_str(v) if v is not None else "")

        cleaned_data.append(out)

    def sort_key(r):
        pd = r[4]
        if isinstance(pd, datetime):
            pd = pd.date()
        elif not isinstance(pd, date):
            pd = date.min
        return (pd, (r[5] or "").lower())

    cleaned_data.sort(key=sort_key)

    out_name = "Clean_All"
    if out_name in wb.sheetnames:
        del wb[out_name]
    out_ws = wb.create_sheet(out_name, 1)

    for col, h in enumerate(CLEAN_HEADERS, 1):
        c = out_ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top", wrap_text=True)

    for ri, row in enumerate(cleaned_data, start=2):
        for ci, val in enumerate(row, start=1):
            cell = out_ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_column_widths(out_ws)

    msg_enter = ensure_enter_sales_sheet(wb)

    wb.save(src)
    print(f"Wrote {len(cleaned_data)} rows to '{out_name}'. {msg_enter}")
    print(f"Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
